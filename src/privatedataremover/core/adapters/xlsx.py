"""Excel (.xlsx) DocumentAdapter — cells, hidden sheets/rows/cols, export."""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Iterator, Sequence

from privatedataremover.core.adapters.base import (
    BBox,
    DocumentAdapter,
    DocumentUnit,
    ExtractedSpan,
    MaskRegion,
)
from privatedataremover.core.export_utils import file_sha256


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "Excel 지원에는 openpyxl이 필요합니다. pip install openpyxl"
        ) from exc
    return openpyxl


class XlsxAdapter(DocumentAdapter):
    """Excel workbook as document units (one sheet = one unit)."""

    format_id = "xlsx"

    def __init__(self) -> None:
        self._path: Path | None = None
        self._wb = None
        self._sheet_names: list[str] = []
        self._original_sha256: str | None = None

    @property
    def unit_count(self) -> int:
        return len(self._sheet_names)

    def open(self, path: Path) -> None:
        openpyxl = _require_openpyxl()
        path = Path(path)
        if not path.is_file():
            raise FileNotFoundError(path)
        self.close()
        self._wb = openpyxl.load_workbook(path, data_only=False)
        self._path = path
        self._sheet_names = list(self._wb.sheetnames)
        self._original_sha256 = file_sha256(path)

    def close(self) -> None:
        if self._wb is not None:
            self._wb.close()
        self._wb = None
        self._path = None
        self._sheet_names = []
        self._original_sha256 = None

    def assert_original_untouched(self) -> None:
        if self._path and self._original_sha256:
            if file_sha256(self._path) != self._original_sha256:
                raise RuntimeError(f"원본 Excel이 변경되었습니다: {self._path}")

    def iter_units(self) -> Iterator[DocumentUnit]:
        self._require()
        for idx, name in enumerate(self._sheet_names):
            ws = self._wb[name]  # type: ignore[index]
            state = getattr(ws, "sheet_state", "visible") or "visible"
            hidden = state in ("hidden", "veryHidden")
            hidden_rows = [
                r for r, dim in ws.row_dimensions.items() if getattr(dim, "hidden", False)
            ]
            hidden_cols = [
                c
                for c, dim in ws.column_dimensions.items()
                if getattr(dim, "hidden", False)
            ]
            # Synthetic size for relative/pattern (columns x rows used)
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            yield DocumentUnit(
                index=idx,
                label=f"{'[숨김] ' if hidden else ''}{name}",
                width=float(max_col),
                height=float(max_row),
                meta={
                    "sheet_name": name,
                    "hidden": hidden,
                    "very_hidden": state == "veryHidden",
                    "sheet_state": state,
                    "hidden_rows": hidden_rows[:50],
                    "hidden_cols": list(hidden_cols)[:50],
                    "hidden_row_count": len(hidden_rows),
                    "hidden_col_count": len(hidden_cols),
                },
            )

    def extract_spans(self, unit_index: int) -> Sequence[ExtractedSpan]:
        self._require()
        ws = self._sheet(unit_index)
        spans: list[ExtractedSpan] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text:
                    continue
                # 0-based col/row in bbox for masking
                col0 = (cell.column or 1) - 1
                row0 = (cell.row or 1) - 1
                # Prefix hidden markers for visibility in analysis UI
                prefix = ""
                if ws.row_dimensions[cell.row].hidden:
                    prefix = "[숨긴행] "
                col_letter = cell.column_letter
                if (
                    col_letter in ws.column_dimensions
                    and ws.column_dimensions[col_letter].hidden
                ):
                    prefix = "[숨긴열] " + prefix
                spans.append(
                    ExtractedSpan(
                        unit_index=unit_index,
                        text=f"{prefix}{text}",
                        bbox=BBox(float(col0), float(row0), float(col0 + 1), float(row0 + 1)),
                        from_ocr=False,
                    )
                )
        return spans

    def render_unit_preview(self, unit_index: int, scale: float = 1.0) -> bytes:
        """Render a simple sheet grid preview with Pillow."""
        from PIL import Image, ImageDraw, ImageFont

        self._require()
        ws = self._sheet(unit_index)
        unit = list(self.iter_units())[unit_index]
        max_rows = min(ws.max_row or 1, 40)
        max_cols = min(ws.max_column or 1, 12)
        cell_w, cell_h = int(90 * scale), int(22 * scale)
        pad = 8
        header_h = 36
        width = pad * 2 + cell_w * max_cols
        height = pad * 2 + header_h + cell_h * max_rows
        img = Image.new("RGB", (width, height), (250, 250, 248))
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.load_default()
        except Exception:  # noqa: BLE001
            font = None

        title = unit.label
        flags = []
        if unit.meta.get("hidden"):
            flags.append("숨긴 시트")
        if unit.meta.get("very_hidden"):
            flags.append("veryHidden")
        if unit.meta.get("hidden_row_count"):
            flags.append(f"숨긴행 {unit.meta['hidden_row_count']}")
        if unit.meta.get("hidden_col_count"):
            flags.append(f"숨긴열 {unit.meta['hidden_col_count']}")
        subtitle = " · ".join(flags) if flags else "표시 시트"
        draw.text((pad, 6), title, fill=(20, 20, 20), font=font)
        draw.text((pad, 20), subtitle, fill=(120, 40, 40) if flags else (80, 80, 80), font=font)

        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                x0 = pad + (c - 1) * cell_w
                y0 = pad + header_h + (r - 1) * cell_h
                x1, y1 = x0 + cell_w, y0 + cell_h
                row_hidden = ws.row_dimensions[r].hidden
                col_letter = ws.cell(1, c).column_letter
                col_hidden = (
                    col_letter in ws.column_dimensions
                    and ws.column_dimensions[col_letter].hidden
                )
                fill = (230, 230, 230) if (row_hidden or col_hidden) else (255, 255, 255)
                draw.rectangle([x0, y0, x1, y1], outline=(180, 180, 180), fill=fill)
                value = ws.cell(r, c).value
                if value is not None:
                    text = str(value)
                    if len(text) > 14:
                        text = text[:13] + "…"
                    draw.text((x0 + 3, y0 + 4), text, fill=(0, 0, 0), font=font)

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    def export_safe(
        self,
        dest: Path,
        masks: Sequence[MaskRegion],
        *,
        text_remove: bool = True,
        draw_black_boxes: bool = True,
        replacement: str = "***",
    ) -> None:
        """Write a copy with masked cells cleared or replaced."""
        openpyxl = _require_openpyxl()
        self._require()
        dest = Path(dest)
        if self._path and dest.resolve() == self._path.resolve():
            raise ValueError("원본 파일과 같은 경로로는 저장할 수 없습니다.")

        # Reload fresh copy from disk so in-memory viewer's book is not mutated.
        wb = openpyxl.load_workbook(self._path)
        try:
            for mask in masks:
                if mask.unit_index < 0 or mask.unit_index >= len(self._sheet_names):
                    continue
                name = self._sheet_names[mask.unit_index]
                ws = wb[name]
                # bbox encodes one or more cells (inclusive 0-based)
                c0 = int(mask.bbox.x0)
                r0 = int(mask.bbox.y0)
                c1 = max(c0, int(mask.bbox.x1) - 1)
                r1 = max(r0, int(mask.bbox.y1) - 1)
                for row in range(r0 + 1, r1 + 2):
                    for col in range(c0 + 1, c1 + 2):
                        cell = ws.cell(row=row, column=col)
                        if text_remove:
                            cell.value = None if replacement == "" else replacement
                        if draw_black_boxes:
                            from openpyxl.styles import PatternFill

                            cell.fill = PatternFill("solid", fgColor="000000")
            wb.save(dest)
        finally:
            wb.close()
        self.assert_original_untouched()

    def export_rasterized(
        self,
        dest: Path,
        masks: Sequence[MaskRegion],
        *,
        dpi: int = 200,
    ) -> None:
        """Export each (masked) sheet preview into a multi-page PDF."""
        import fitz

        self._require()
        dest = Path(dest)
        # Apply masks on a temp xlsx then preview — simpler: preview current + black cells on image
        # For correctness, export_safe to temp then reopen.
        import tempfile

        with tempfile.TemporaryDirectory(prefix="pdr_xlsx_") as td:
            tmp = Path(td) / "masked.xlsx"
            self.export_safe(tmp, masks)
            other = XlsxAdapter()
            other.open(tmp)
            try:
                out = fitz.open()
                scale = max(0.8, dpi / 150.0)
                for i, _ in enumerate(other.iter_units()):
                    png = other.render_unit_preview(i, scale=scale)
                    pix = fitz.Pixmap(png)
                    page = out.new_page(width=pix.width, height=pix.height)
                    page.insert_image(page.rect, pixmap=pix)
                out.save(dest)
                out.close()
            finally:
                other.close()
        self.assert_original_untouched()

    def list_hidden_sheet_summary(self) -> list[dict]:
        return [
            {
                "index": u.index,
                "name": u.meta.get("sheet_name"),
                "hidden": u.meta.get("hidden"),
                "very_hidden": u.meta.get("very_hidden"),
                "hidden_rows": u.meta.get("hidden_row_count"),
                "hidden_cols": u.meta.get("hidden_col_count"),
            }
            for u in self.iter_units()
            if u.meta.get("hidden")
            or u.meta.get("hidden_row_count")
            or u.meta.get("hidden_col_count")
        ]

    def _sheet(self, unit_index: int):
        self._require()
        if unit_index < 0 or unit_index >= len(self._sheet_names):
            raise IndexError(unit_index)
        return self._wb[self._sheet_names[unit_index]]  # type: ignore[index]

    def _require(self) -> None:
        if self._wb is None:
            raise RuntimeError("Excel 파일이 열려 있지 않습니다.")
