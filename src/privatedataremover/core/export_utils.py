"""Export helpers: residual text check and temp cleanup."""

from __future__ import annotations

import hashlib
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence


@dataclass(frozen=True)
class ResidualHit:
    text: str
    page_index: int | None = None


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def find_residual_texts(pdf_path: Path, forbidden: Sequence[str]) -> list[ResidualHit]:
    """Return forbidden strings still extractable from a PDF."""
    import fitz

    cleaned = [t.strip() for t in forbidden if t and t.strip()]
    if not cleaned:
        return []

    hits: list[ResidualHit] = []
    doc = fitz.open(pdf_path)
    try:
        for i, page in enumerate(doc):
            page_text = page.get_text("text") or ""
            for needle in cleaned:
                if needle in page_text:
                    hits.append(ResidualHit(text=needle, page_index=i))
    finally:
        doc.close()
    return hits


def find_residual_in_xlsx(xlsx_path: Path, forbidden: Sequence[str]) -> list[ResidualHit]:
    """Return forbidden strings still present in workbook cell values."""
    try:
        import openpyxl
    except ImportError:
        return []

    cleaned = [t.strip() for t in forbidden if t and t.strip()]
    if not cleaned:
        return []

    hits: list[ResidualHit] = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        for sheet_index, name in enumerate(wb.sheetnames):
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    value = str(cell.value)
                    for needle in cleaned:
                        if needle in value:
                            hits.append(ResidualHit(text=needle, page_index=sheet_index))
    finally:
        wb.close()
    return hits


class TempWorkspace:
    """Temporary directory cleaned up explicitly or on context exit."""

    def __init__(self, prefix: str = "pdr_") -> None:
        self._dir = Path(tempfile.mkdtemp(prefix=prefix))

    @property
    def path(self) -> Path:
        return self._dir

    def tempfile(self, name: str) -> Path:
        return self._dir / name

    def cleanup(self) -> None:
        if self._dir.exists():
            shutil.rmtree(self._dir, ignore_errors=True)

    def __enter__(self) -> TempWorkspace:
        return self

    def __exit__(self, *exc) -> None:
        self.cleanup()
